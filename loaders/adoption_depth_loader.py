from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook

from config.deployment import repository_writes_enabled
from helpers.atomic_io import atomic_write_csv

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEPTH_PATH = PROJECT_ROOT / "data" / "adoption_ai_supplement_2026.csv"
BTOS_AI_SUPPLEMENT_URL = "https://www.census.gov/hfp/btos/downloads/AI_Supplement_Table_2026.xlsx"
BTOS_REQUEST_HEADERS = {
    "User-Agent": "AI Macro research client",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream;q=0.9,*/*;q=0.8",
    "Referer": "https://www.census.gov/hfp/btos/data_downloads",
}
REFERENCE_START = "2025-11-17"
REFERENCE_END = "2026-02-08"

SOURCE_COLUMNS = (
    "Scope",
    "Question ID",
    "Question",
    "Answer ID",
    "Answer",
    "Measure 1 Header",
    "Measure 1",
    "Measure 1 Standard Error Header",
    "Measure 1 Standard Error",
    "Measure 1 Status",
    "Measure 2 Header",
    "Measure 2",
    "Measure 2 Standard Error Header",
    "Measure 2 Standard Error",
    "Measure 2 Status",
    "Source Sheet",
    "Source Row",
    "Reference Start",
    "Reference End",
)


def _load_report(source_mode: str, error: str = "") -> dict:
    return {
        "source_mode": source_mode,
        "error": str(error or ""),
        "dataset_kind": "fixed_published_supplement",
        "retained_path": str(DEPTH_PATH.relative_to(PROJECT_ROOT)),
        "reference_start": REFERENCE_START,
        "reference_end": REFERENCE_END,
    }


def _clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _key(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _clean(value).casefold()).strip()


def _integer(value) -> int | None:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return None
    number = float(numeric)
    return int(number) if number.is_integer() else None


def _percentage(value) -> float:
    if value is None:
        return np.nan
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return np.nan
        if text.endswith("%"):
            numeric = pd.to_numeric(text[:-1], errors="coerce")
            return float(numeric) if pd.notna(numeric) else np.nan
        numeric = pd.to_numeric(text, errors="coerce")
        return float(numeric) if pd.notna(numeric) else np.nan
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return np.nan
    number = float(numeric)
    return number * 100.0 if 0 <= number <= 1 else number


def _status(value) -> str:
    if pd.notna(_percentage(value)):
        return "observed"
    return "missing" if not _clean(value) else "suppressed"


def _sheet(workbook, name: str):
    target = _key(name)
    for sheet_name in workbook.sheetnames:
        if _key(sheet_name) == target:
            return workbook[sheet_name]
    available = ", ".join(workbook.sheetnames)
    raise ValueError(f"BTOS AI supplement missing sheet '{name}'; sheets: {available}")


def _header_contract(sheet) -> tuple[int, list[str], dict[str, int]]:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        labels = [_clean(value) for value in row]
        keys = {_key(value): index for index, value in enumerate(labels) if value}
        if "question id" in keys and "answer id" in keys and "answer" in keys:
            return row_number, labels, keys
    raise ValueError(f"BTOS AI supplement sheet '{sheet.title}' has no recognized header row")


def _column(keys: dict[str, int], label: str, *aliases: str) -> int:
    candidates = (label, *aliases)
    for candidate in candidates:
        target = _key(candidate)
        if target in keys:
            return keys[target]

    for candidate in candidates:
        target = _key(candidate)
        matches = [(key, index) for key, index in keys.items() if key.startswith(target)]
        if len(matches) == 1:
            return matches[0][1]
        if len(matches) > 1:
            available = ", ".join(sorted(keys))
            raise ValueError(
                f"BTOS AI supplement column '{candidate}' is ambiguous; "
                f"matches={[key for key, _ in matches]}; columns={available}"
            )

    expected = ", ".join(repr(candidate) for candidate in candidates)
    available = ", ".join(sorted(keys))
    raise ValueError(f"BTOS AI supplement missing column; expected one of {expected}; columns={available}")


def _measure_columns(
    labels: list[str],
    keys: dict[str, int],
    first: str,
    second: str,
) -> tuple[tuple[int, str], tuple[int, str]]:
    first_col = _column(keys, first)
    second_col = _column(keys, second)
    return (first_col, labels[first_col]), (second_col, labels[second_col])


def _data_rows(sheet) -> tuple[int, list[tuple], list[str], dict[str, int]]:
    header_row, labels, keys = _header_contract(sheet)
    rows = list(sheet.iter_rows(min_row=header_row + 1, values_only=True))
    return header_row, rows, labels, keys


def parse_btos_ai_supplement_workbook(content: bytes) -> pd.DataFrame:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    response_sheet = _sheet(workbook, "National Response Estimates")
    error_sheet = _sheet(workbook, "National Standard Errors")

    response_header, response_rows, response_labels, response_keys = _data_rows(response_sheet)
    _, error_rows, error_labels, error_keys = _data_rows(error_sheet)

    scope_col = _column(response_keys, "Scope")
    qid_col = _column(response_keys, "Question ID")
    question_col = _column(response_keys, "Question")
    answer_id_col = _column(response_keys, "Answer ID")
    answer_col = _column(response_keys, "Answer")
    measure_1, measure_2 = _measure_columns(response_labels, response_keys, "Estimate", "Estimate - Yes")

    error_scope_col = _column(error_keys, "Scope")
    error_qid_col = _column(error_keys, "Question ID")
    error_answer_id_col = _column(error_keys, "Answer ID")
    error_measure_1_col = _column(error_keys, "SE", "Standard Error")
    error_measure_2_col = _column(error_keys, "SE - Yes", "Standard Error - Yes")
    error_measure_1 = (error_measure_1_col, error_labels[error_measure_1_col])
    error_measure_2 = (error_measure_2_col, error_labels[error_measure_2_col])

    errors: dict[tuple[int, int, int], tuple[float, float]] = {}
    for row in error_rows:
        scope = _integer(row[error_scope_col])
        question_id = _integer(row[error_qid_col])
        answer_id = _integer(row[error_answer_id_col])
        if scope is None or question_id is None or answer_id is None:
            continue
        errors[(scope, question_id, answer_id)] = (
            _percentage(row[error_measure_1[0]]),
            _percentage(row[error_measure_2[0]]),
        )

    records: list[dict] = []
    for offset, row in enumerate(response_rows, start=1):
        scope = _integer(row[scope_col])
        question_id = _integer(row[qid_col])
        answer_id = _integer(row[answer_id_col])
        if scope is None or question_id is None or answer_id is None:
            continue

        raw_1 = row[measure_1[0]]
        raw_2 = row[measure_2[0]]
        error_1, error_2 = errors.get((scope, question_id, answer_id), (np.nan, np.nan))
        records.append({
            "Scope": scope,
            "Question ID": question_id,
            "Question": _clean(row[question_col]),
            "Answer ID": answer_id,
            "Answer": _clean(row[answer_col]),
            "Measure 1 Header": measure_1[1],
            "Measure 1": _percentage(raw_1),
            "Measure 1 Standard Error Header": error_measure_1[1],
            "Measure 1 Standard Error": error_1,
            "Measure 1 Status": _status(raw_1),
            "Measure 2 Header": measure_2[1],
            "Measure 2": _percentage(raw_2),
            "Measure 2 Standard Error Header": error_measure_2[1],
            "Measure 2 Standard Error": error_2,
            "Measure 2 Status": _status(raw_2),
            "Source Sheet": response_sheet.title,
            "Source Row": response_header + offset,
            "Reference Start": REFERENCE_START,
            "Reference End": REFERENCE_END,
        })

    output = pd.DataFrame(records, columns=SOURCE_COLUMNS)
    if output.empty:
        raise ValueError("BTOS AI supplement workbook produced no national response rows")

    duplicates = output.duplicated(["Scope", "Question ID", "Answer ID"], keep=False)
    if duplicates.any():
        keys = output.loc[duplicates, ["Scope", "Question ID", "Answer ID"]].drop_duplicates().to_dict("records")
        raise ValueError(f"BTOS AI supplement contains duplicate national response keys: {keys[:8]}")
    return output.sort_values(["Scope", "Question ID", "Answer ID"], kind="stable").reset_index(drop=True)


def _load_retained() -> pd.DataFrame:
    if not DEPTH_PATH.exists() or not DEPTH_PATH.stat().st_size:
        return pd.DataFrame(columns=SOURCE_COLUMNS)
    frame = pd.read_csv(DEPTH_PATH)
    if not set(SOURCE_COLUMNS).issubset(frame.columns):
        return pd.DataFrame(columns=SOURCE_COLUMNS)
    for column in ["Scope", "Question ID", "Answer ID", "Source Row"]:
        frame[column] = pd.to_numeric(frame[column], errors="coerce").astype("Int64")
    for column in ["Measure 1", "Measure 1 Standard Error", "Measure 2", "Measure 2 Standard Error"]:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame[list(SOURCE_COLUMNS)].reset_index(drop=True)


def persist_adoption_depth_source(frame: pd.DataFrame) -> None:
    if not repository_writes_enabled() or not isinstance(frame, pd.DataFrame) or frame.empty:
        return
    atomic_write_csv(frame[list(SOURCE_COLUMNS)], DEPTH_PATH)


def load_adoption_depth(*, force_refresh: bool = False, allow_live: bool = False) -> dict:
    retained = _load_retained()
    table = retained
    report = (
        _load_report("retained_official")
        if not retained.empty
        else _load_report(
            "unavailable",
            f"Required retained Census AI supplement is missing or invalid: "
            f"{DEPTH_PATH.relative_to(PROJECT_ROOT)}",
        )
    )

    if force_refresh and allow_live:
        try:
            response = requests.get(BTOS_AI_SUPPLEMENT_URL, timeout=60, headers=BTOS_REQUEST_HEADERS)
            response.raise_for_status()
            table = parse_btos_ai_supplement_workbook(response.content)
            report = _load_report("live_candidate")
        except Exception as exc:
            table = retained
            report = _load_report(
                "retained_fallback" if not retained.empty else "unavailable",
                f"{type(exc).__name__}: {exc}",
            )

    return {
        "table": table,
        "retained_table": retained,
        "load_report": report,
        "source": "U.S. Census BTOS AI Supplement 2026",
    }
